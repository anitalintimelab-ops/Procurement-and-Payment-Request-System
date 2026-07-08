import streamlit as st
import pandas as pd
import datetime
import os
import io
from openpyxl.styles import Border, Side, Alignment
import openpyxl

# ... (省略前方與之前一致的工具函式，請保持原樣) ...

# ★ 強制繞過合併儲存格填寫 Excel 的核心函數
def safe_write_and_style(ws_obj, r_idx, c_idx, val=None, border=None, alignment=None):
    cell_obj = ws_obj.cell(row=r_idx, column=c_idx)
    # 如果是合併儲存格，轉向左上角主格
    if isinstance(cell_obj, openpyxl.cell.cell.MergedCell) or type(cell_obj).__name__ == 'MergedCell':
        for m_range in ws_obj.merged_cells.ranges:
            if cell_obj.coordinate in m_range:
                cell_obj = ws_obj.cell(row=m_range.min_row, column=m_range.min_col)
                break
    
    if val is not None: cell_obj.value = val
    if border is not None: cell_obj.border = border
    if alignment is not None: cell_obj.alignment = alignment

# ... (中間的系統邏輯保持不變) ...

# ★ 修正後的報表產生區塊
elif menu == "5. 產出本期支出報表":
    # ... (前面的勾選與參數設定保持不變) ...
    if st.button("🚀 產生 Excel 支出報表"):
        # ... (讀取 template 邏輯) ...
        wb = openpyxl.load_workbook(TEMPLATE_FILE)
        ws = wb.active
        # ... (使用 safe_write_and_style 寫入資料，不再產生 MergedCell 錯誤) ...
        # 餘額填寫邏輯
        balance_after = float(balance_before) - sum_transfer 
        # 這裡根據您的截圖，第8欄(Column H)是合計與餘額的目標欄位，請確認您的範本 H 欄是否對應
        safe_write_and_style(ws, summary_row, 8, balance_after) 
        wb.save(output)

# ★ 新增的專案/廠商管理介面
elif menu == "7. 專案 / 廠商資料庫":
    st.subheader("🗂️ 專案 / 廠商資料庫管理")
    t1, t2 = st.tabs(["📂 專案資料庫", "🏢 廠商資料庫"])
    with t1:
        p_db = load_projects()
        edited_p = st.data_editor(p_db, num_rows="dynamic", use_container_width=True)
        if st.button("💾 儲存專案變更"): save_projects(edited_p); st.success("已更新"); st.rerun()
    with t2:
        v_db = load_vendors()
        edited_v = st.data_editor(v_db, num_rows="dynamic", use_container_width=True)
        if st.button("💾 儲存廠商變更"): save_vendors(edited_v); st.success("已更新"); st.rerun()
